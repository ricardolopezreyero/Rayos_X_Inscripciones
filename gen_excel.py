"""
Genera Método_Comercial_SuperLeads_2026.xlsx
Con 7 hojas: Pipeline, DoD, Expansión, Ritmo, Objeciones, Scorecard, Equipo
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Colores SuperLeads ─────────────────────────────────────────────────────────
NAVY_DEEP  = "001240"
NAVY_MID   = "001A5E"
NAVY       = "002582"
GREEN      = "56EF9F"
GREEN_D    = "2BC878"
BLUE_MID   = "2A89FB"
WHITE      = "FFFFFF"
MUTED      = "7A9FD4"
DARK_TEXT  = "E0EAFF"
LIGHT_ROW  = "02174A"   # fila impar
ALT_ROW    = "011035"   # fila par

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="bottom"):
    s = Side(style="thin", color="0D2A6E")
    kw = {}
    for side in sides.split():
        kw[side] = s
    return Border(**kw)

def thick_left():
    return Border(left=Side(style="medium", color=GREEN))

def setup_sheet(ws, tab_color):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False


# ── ROW HELPERS ───────────────────────────────────────────────────────────────

def header_row(ws, row, cols, col_start=1, bg=NAVY_MID, fg=GREEN, size=9):
    for i, text in enumerate(cols):
        c = ws.cell(row=row, column=col_start + i, value=text)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=size)
        c.alignment = align(h="left", v="center", wrap=False)
        c.border = thin_border("bottom")

def data_row(ws, row, values, col_start=1, bg=None, fg=WHITE):
    bg = bg or (LIGHT_ROW if row % 2 == 1 else ALT_ROW)
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=str(val) if val else "")
        c.fill = fill(bg)
        c.font = font(bold=False, color=fg, size=9)
        c.alignment = align(h="left", v="top", wrap=True)
        c.border = thin_border("bottom")

def section_title(ws, row, text, ncols, col_start=1):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_start + ncols - 1)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(NAVY_DEEP)
    c.font = font(bold=True, color=GREEN, size=11)
    c.alignment = align(h="left", v="center", wrap=False)
    return row

def spacer(ws, row, ncols, col_start=1):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_start + ncols - 1)
    c = ws.cell(row=row, column=col_start, value="")
    c.fill = fill(NAVY_DEEP)
    ws.row_dimensions[row].height = 6

def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def set_height(ws, row, h):
    ws.row_dimensions[row].height = h

def green_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=True, color=GREEN, size=9)
    return c


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1: PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

STAGES = [
    {
        "num": "01", "name": "Lead", "resp": "Setter / Vendedor",
        "quien": "Inbound: redes sociales, WhatsApp, Meta Ads, Google Ads. Outbound: bases propias, referidos, CIMEDU.",
        "objetivo": "Lograr primera respuesta real. Pasar a Contactado lo antes posible.",
        "tiempo_max": "1 día", "descuido_max": "3 horas", "max_negocios": "Sin límite",
        "actividades": "Cadencia: Día 0: WA + llamada · Día 1: WA corto · Día 3: ángulo diferente · Día 7: último intento · Día 14: cierre formal",
        "perdidos": "No contesta / inaccesible · Fuente de baja calidad · Duplicado · No cumple el perfil ideal",
        "entregables": "—",
        "recibimos": "Nombre + WhatsApp + correo · Fuente de origen",
        "automatizaciones": "IA conversacional fuera de horario · Cadencia automática WA/email días 0-1-3-7-14 · Asignación al owner en CRM",
        "costo": "~$750–800 MXN por agendamiento (Meta Ads + Google). Presupuesto: $27,000/mes"
    },
    {
        "num": "02", "name": "Contactado", "resp": "Setter",
        "quien": "Lead que ya respondió al menos una vez.",
        "objetivo": "Confirmar que vale la pena seguir. Agendar el Diagnóstico Rayos X.",
        "tiempo_max": "3 días", "descuido_max": "5 horas", "max_negocios": "Sin límite",
        "actividades": "2 intentos de reagendado con horarios concretos · 1 llamada de rescate · Mensaje de cierre si no responde",
        "perdidos": "Solo quiere campañas sin sistema · Nadie decide ni ejecuta · No comparte datos básicos",
        "entregables": "Mensaje de presentación · Link para agendar el Diagnóstico · Video corto (opcional)",
        "recibimos": "Rol en la institución · Dolor principal en admisiones · Urgencia por ciclo",
        "automatizaciones": "Alerta si lead no avanza en 5 h · Secuencia de follow-up automática",
        "costo": "—"
    },
    {
        "num": "03", "name": "Calificado", "resp": "Setter",
        "quien": "Contacto con dolor real en admisiones, capacidad de pago y poder de decisión.",
        "objetivo": "Que confirme la cita cumpliendo los 3 requisitos: ① Evento enviado ② Confirmó por WA ③ Mensaje 'qué preparar' enviado.",
        "tiempo_max": "5 días", "descuido_max": "8 horas", "max_negocios": "Sin límite",
        "actividades": "2 intentos de confirmar la cita · Si no confirma en 48 h: reagendar o marcar perdido",
        "perdidos": "Sin presupuesto claro · No tiene autoridad para decidir · No es el momento",
        "entregables": "Evento en Google Calendar · Mensaje 'Qué preparar': alumnos · meta · ticket · capacidad · etapa del ciclo",
        "recibimos": "Alumnos actuales · Meta del ciclo · Ticket promedio anual · Capacidad máxima · Etapa del ciclo actual",
        "automatizaciones": "Mensaje 'qué preparar' automático · Evento de Google Calendar generado",
        "costo": "—"
    },
    {
        "num": "04", "name": "Cita Agendada", "resp": "Setter + Vendedor",
        "quien": "Calificado que acordó fecha y hora para el Diagnóstico Rayos X.",
        "objetivo": "Que sí asista. Nada más importa en esta etapa.",
        "tiempo_max": "Hasta la cita (máx. 7 días)", "descuido_max": "24 horas", "max_negocios": "25 activas por consultor",
        "actividades": "WA a los 10 min post no-show · 2 horarios alternativos el mismo día · Cadencia 1/3/7 si no responde",
        "perdidos": "No-show sin reagendado · Perdió el interés · Competidor fue más rápido",
        "entregables": "Recordatorio WA 24 h antes · Recordatorio WA 2 h antes con link",
        "recibimos": "Confirmación de asistencia por WhatsApp.",
        "automatizaciones": "Recordatorio WA 24 h y 2 h antes · Alerta si cita no fue confirmada en 48 h",
        "costo": "—"
    },
    {
        "num": "05", "name": "Cita Asistida", "resp": "Vendedor",
        "quien": "Prospecto que sí se presentó a la cita.",
        "objetivo": "Salir con 3 cuellos documentados, impacto económico calculado y 2ª reunión agendada en el acto.",
        "tiempo_max": "Mismo día", "descuido_max": "—", "max_negocios": "—",
        "actividades": "Intentar reagendar x1 si no hubo diagnóstico completo",
        "perdidos": "No asistió y no reagendó",
        "entregables": "—",
        "recibimos": "Información completa del proceso de admisión actual.",
        "automatizaciones": "Cambio de etapa automático al registrar asistencia · Tarea 'Preparar propuesta' creada",
        "costo": "—"
    },
    {
        "num": "06", "name": "Diagnóstico", "resp": "Vendedor",
        "quien": "Prospecto con diagnóstico completo: 3 cuellos e impacto económico calculado.",
        "objetivo": "Propuesta que conecte exactamente: cuello → solución → inversión vs. dinero no capturado.",
        "tiempo_max": "7 días para entregar propuesta", "descuido_max": "48 horas", "max_negocios": "—",
        "actividades": "Enviar propuesta 2 veces · Loom personalizado · Radiografía extendida · Fecha límite explícita",
        "perdidos": "No quiere segunda reunión · No acepta los cuellos detectados",
        "entregables": "PDF del Diagnóstico SuperLeads: 3 cuellos detectados · Nivel actual del sistema · Impacto económico estimado · Nivel recomendado",
        "recibimos": "Validación de los cuellos detectados · Apertura para la propuesta",
        "automatizaciones": "Generación del PDF de Diagnóstico (App Rayos X) · Alerta: 'Enviar propuesta en 7 días'",
        "costo": "App Rayos X: $8,000 MXN (pago único)"
    },
    {
        "num": "07", "name": "Propuesta Entregada", "resp": "Vendedor",
        "quien": "Prospecto que recibió la propuesta formal y está en revisión activa.",
        "objetivo": "Llegar a decisión con fecha. Nunca dejar propuesta abierta sin siguiente paso.",
        "tiempo_max": "7 días para obtener decisión", "descuido_max": "48 horas", "max_negocios": "20 activas",
        "actividades": "Caso de éxito similar · Recordatorio del impacto económico · Fecha límite de decisión",
        "perdidos": "Precio alto · Ya trabajan con alguien · El que decide no está disponible",
        "entregables": "Propuesta formal: Diagnóstico recap · Cuello → solución → inversión · Siguiente paso con fecha",
        "recibimos": "Retroalimentación sobre la propuesta · Objeciones explícitas",
        "automatizaciones": "Alerta stale si propuesta no se mueve en 48 h · Notificación interna si supera 7 días",
        "costo": "Incluido en Dirección Estratégica ($6,000–$10,000/mes)"
    },
    {
        "num": "08", "name": "Negociación", "resp": "Vendedor + Director (cierre)",
        "quien": "Interesado evaluando condiciones finales antes de decidir.",
        "objetivo": "Cerrar con firma y pago. Sin perder momentum después del sí.",
        "tiempo_max": "14 días activos", "descuido_max": "72 horas", "max_negocios": "15 activas",
        "actividades": "Resolver objeción desde impacto, no desde precio · Video 'Qué es SuperLeads' · Último activo antes de cerrar",
        "perdidos": "Precio fuera de rango · Sin urgencia real · Hallazgo estructural bloqueante · Eligieron a otro",
        "entregables": "Caso de éxito similar · Loom personalizado · Radiografía extendida · Resumen 3 cuellos exactos",
        "recibimos": "Condiciones finales · Quién firma y quién paga",
        "automatizaciones": "Alerta stale cada 72 h · Webhook al equipo si llega a 14 días sin avance",
        "costo": "—"
    },
    {
        "num": "09", "name": "Aceptada", "resp": "Administración + Vendedor",
        "quien": "Prospecto que dijo sí, firmó y está listo para pagar.",
        "objetivo": "Ejecutar el cobro rápido y sin fricción.",
        "tiempo_max": "3 días para cobrar", "descuido_max": "24 horas", "max_negocios": "Sin límite",
        "actividades": "Link de pago enviado 2 veces · Llamada de cierre si no se mueve",
        "perdidos": "Decisión revertida · Cambio de prioridades interno",
        "entregables": "Contrato para firma · Link de pago (Stripe)",
        "recibimos": "Firma del contrato.",
        "automatizaciones": "Link de pago Stripe generado · Notificación a Administración",
        "costo": "—"
    },
    {
        "num": "10", "name": "Pagada", "resp": "Conexión",
        "quien": "Cliente con pago confirmado, listo para implementación.",
        "objetivo": "Handoff impecable al equipo de implementación.",
        "tiempo_max": "24 h para arrancar handoff", "descuido_max": "24 horas", "max_negocios": "Sin límite",
        "actividades": "—",
        "perdidos": "—",
        "entregables": "Handoff completo · Grupo de trabajo creado · Kickoff agendado",
        "recibimos": "Pago confirmado en Stripe.",
        "automatizaciones": "Webhook al equipo de implementación · Notificación al equipo · Grupo de onboarding creado",
        "costo": "CRM Básico: $12,000/mes · CRM Avanzado: $25,000/mes (+ IVA)"
    },
    {
        "num": "—", "name": "Perdido", "resp": "Vendedor",
        "quien": "Oportunidad que no avanzó en cualquier etapa.",
        "objetivo": "Registrar la razón exacta. Sin dato no hay mejora.",
        "tiempo_max": "—", "descuido_max": "—", "max_negocios": "—",
        "actividades": "—",
        "perdidos": "Ver tipificaciones por etapa en el pipeline.",
        "entregables": "—",
        "recibimos": "Razón concreta de cierre.",
        "automatizaciones": "Cierre con tipificación obligatoria en CRM · Pipeline actualizado automáticamente",
        "costo": "—"
    }
]


def build_pipeline(wb):
    ws = wb.create_sheet("🔄 Pipeline")
    setup_sheet(ws, GREEN)

    ws["A1"] = "PIPELINE COMERCIAL — MÉTODO SUPERLEADS 2026"
    ws.merge_cells("A1:L1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "11 etapas (10 activas + Perdido). Cada etapa tiene dueño, objetivo y tiempo máximo de descuido."
    ws.merge_cells("A2:L2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(bold=False, size=9, color=MUTED)
    ws["A2"].alignment = align(h="left")

    spacer(ws, 3, 12)

    COLS = ["#", "Etapa", "Responsable", "¿Quiénes llegan?", "Objetivo de la etapa",
            "Tiempo máx. en etapa", "Tiempo de descuido", "Máx. negocios activos",
            "Actividades antes de marcar perdido", "Tipificaciones de perdidos",
            "Automatizaciones", "Costo / Dato económico"]

    header_row(ws, 4, COLS)
    set_height(ws, 4, 22)

    for r, s in enumerate(STAGES, 5):
        bg = LIGHT_ROW if r % 2 == 1 else ALT_ROW
        row_vals = [
            s["num"], s["name"], s["resp"], s["quien"], s["objetivo"],
            s["tiempo_max"], s["descuido_max"], s["max_negocios"],
            s["actividades"], s["perdidos"], s["automatizaciones"], s["costo"]
        ]
        data_row(ws, r, row_vals, bg=bg)
        # Highlight etapa name in green
        nc = ws.cell(row=r, column=2)
        nc.font = font(bold=True, color=GREEN, size=9)
        # Highlight responsable in blue-mid
        rc = ws.cell(row=r, column=3)
        rc.font = font(bold=True, color=BLUE_MID, size=9)
        set_height(ws, r, 60)

    set_widths(ws, [4, 16, 22, 40, 40, 20, 18, 20, 48, 40, 50, 35])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2: DEFINITION OF DONE
# ══════════════════════════════════════════════════════════════════════════════

DOD = [
    {
        "etapa": "Lead", "tiempo": "Mismo día", "resp": "Setter",
        "criterios": "☑ Lead tiene nombre + WhatsApp válido\n☑ Fuente registrada en CRM\n☑ Owner asignado\n☑ Primer mensaje enviado en <3 horas",
        "dato_crm": "Nombre, WA, fuente, owner, fecha primer contacto"
    },
    {
        "etapa": "Contactado", "tiempo": "Máx. 3 días", "resp": "Setter",
        "criterios": "☑ Prospecto respondió al menos 1 vez\n☑ Rol en la institución confirmado\n☑ Existe dolor real en admisiones\n☑ Tiene poder de decisión o lo refiere a quien lo tiene",
        "dato_crm": "Rol, dolor declarado, nivel de autoridad, fuente confirmada"
    },
    {
        "etapa": "Calificado", "tiempo": "Máx. 5 días", "resp": "Setter",
        "criterios": "☑ Los 5 datos para Rayos X confirmados: alumnos · meta ciclo · ticket · capacidad · etapa\n☑ Tiene presupuesto o autoridad para aprobarlo\n☑ Urgencia real detectada (ciclo, meta, presión)",
        "dato_crm": "5 datos de diagnóstico, presupuesto o decisor identificado"
    },
    {
        "etapa": "Cita Agendada", "tiempo": "Antes de la cita", "resp": "Setter",
        "criterios": "☑ Evento en Google Calendar enviado y aceptado\n☑ Confirmación por WhatsApp recibida\n☑ Mensaje 'qué preparar' enviado\n☑ Recordatorio 24h y 2h programados",
        "dato_crm": "Fecha/hora cita, confirmación WA, link enviado"
    },
    {
        "etapa": "Cita Asistida", "tiempo": "Mismo día", "resp": "Vendedor",
        "criterios": "☑ Prospecto asistió y diagnóstico completado\n☑ 3 cuellos de botella documentados\n☑ Impacto económico calculado (brecha × ticket)\n☑ Segunda reunión agendada en el acto",
        "dato_crm": "3 cuellos, impacto económico MXN, fecha 2ª reunión"
    },
    {
        "etapa": "Diagnóstico", "tiempo": "Máx. 7 días", "resp": "Vendedor",
        "criterios": "☑ PDF de Diagnóstico generado y enviado al cliente\n☑ Propuesta estructurada: cuellos → solución → inversión vs. dinero no capturado\n☑ Fecha de decisión acordada explícitamente\n☑ Propuesta revisada por Vendedor",
        "dato_crm": "PDF enviado, propuesta en Drive, fecha decisión registrada"
    },
    {
        "etapa": "Propuesta", "tiempo": "Máx. 7 días", "resp": "Vendedor",
        "criterios": "☑ Cliente leyó la propuesta (confirmado)\n☑ Objeciones principales identificadas y registradas\n☑ Decisor confirmado (quien firma + quien paga)\n☑ Siguiente paso concreto acordado con fecha",
        "dato_crm": "Objeciones registradas, decisor confirmado, fecha decisión"
    },
    {
        "etapa": "Negociación", "tiempo": "Máx. 14 días", "resp": "Vendedor + Director",
        "criterios": "☑ Condiciones finales confirmadas\n☑ Activos de cierre utilizados (caso similar, loom, radiografía)\n☑ Descuento aprobado por Director (si aplica)\n☑ Contrato enviado para firma",
        "dato_crm": "Condiciones acordadas, descuento autorizado, contrato enviado"
    },
    {
        "etapa": "Aceptada", "tiempo": "Máx. 3 días", "resp": "Admin + Vendedor",
        "criterios": "☑ Contrato firmado recibido\n☑ Link de pago Stripe enviado\n☑ Fecha de pago comprometida\n☑ Administración notificada",
        "dato_crm": "Contrato firmado, fecha de pago, notificación a Admin"
    },
    {
        "etapa": "Pagada", "tiempo": "Máx. 24h post-pago", "resp": "Conexión",
        "criterios": "☑ Pago confirmado en Stripe\n☑ Handoff completo preparado (diagnóstico, promesas, riesgos, urgencias)\n☑ Grupo de trabajo creado\n☑ Kickoff agendado en los próximos 5 días hábiles\n☑ Equipo de implementación notificado",
        "dato_crm": "Pago confirmado, handoff en Drive, kickoff agendado"
    }
]

def build_dod(wb):
    ws = wb.create_sheet("✅ Definition of Done")
    setup_sheet(ws, GREEN_D)

    ws["A1"] = "DEFINITION OF DONE — CRITERIOS DE AVANCE POR ETAPA"
    ws.merge_cells("A1:E1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "Regla absoluta: si no se cumplen TODOS los criterios, la oportunidad NO avanza. Sin este estándar el proceso no es escalable."
    ws.merge_cells("A2:E2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(bold=True, size=9, color="FF8080")
    ws["A2"].alignment = align(h="left")

    spacer(ws, 3, 5)
    header_row(ws, 4, ["Etapa", "Criterios obligatorios (TODOS)", "Dato mínimo en CRM", "Responsable", "Tiempo límite"])
    set_height(ws, 4, 22)

    for r, d in enumerate(DOD, 5):
        bg = LIGHT_ROW if r % 2 == 1 else ALT_ROW
        data_row(ws, r, [d["etapa"], d["criterios"], d["dato_crm"], d["resp"], d["tiempo"]], bg=bg)
        ws.cell(row=r, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=r, column=4).font = font(bold=True, color=BLUE_MID, size=9)
        ws.cell(row=r, column=5).font = font(bold=True, color=GREEN, size=9)
        set_height(ws, r, 70)

    set_widths(ws, [18, 60, 42, 22, 20])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3: RUTA DE EXPANSIÓN
# ══════════════════════════════════════════════════════════════════════════════

EXPANSION = [
    {
        "nivel": "Nivel 1 · Entrada", "producto": "CRM SuperLeads Básico", "precio": "$12,000 / mes",
        "cuando": "Siempre primero. Sin base no hay sistema.",
        "quien": "Vendedor", "mrr_acum": "$12,000/mes",
        "senal": "Cualquier institución sin CRM activo o con proceso en Excel/WhatsApp."
    },
    {
        "nivel": "Nivel 1+ · Upgrade base", "producto": "CRM SuperLeads Avanzado", "precio": "$25,000 / mes",
        "cuando": "Cuando el cliente ya usa el Básico y pide más automatización, más campus o más volumen.",
        "quien": "Conexión", "mrr_acum": "$25,000/mes",
        "senal": "Cliente en Básico con >3 meses activo, pide reportes más detallados o tiene >1 campus."
    },
    {
        "nivel": "Nivel 2A · Dirección", "producto": "Dirección Estratégica Básico", "precio": "+ $6,000 / mes",
        "cuando": "En la primera revisión de cuenta (mes 1-2). El cliente ve los huecos y los pide.",
        "quien": "Vendedor / Conexión", "mrr_acum": "$18,000–$31,000/mes",
        "senal": "Cliente con CRM activo que pregunta '¿qué sigue?' o tiene estancada la conversión."
    },
    {
        "nivel": "Nivel 2B · Campañas", "producto": "Campañas Integrales Básico", "precio": "+ $10,000 / mes",
        "cuando": "Cuando el sistema está ordenado pero el volumen de leads es insuficiente.",
        "quien": "Vendedor", "mrr_acum": "$28,000–$41,000/mes",
        "senal": "Cliente con CRM + Dirección activos. Tasa de contactación buena pero pocos leads entrando."
    },
    {
        "nivel": "Nivel 2C · Puntos de contacto", "producto": "Optimización Puntos de Contacto", "precio": "$10,000–$30,000 único",
        "cuando": "Cuando el diagnóstico muestra fricción digital: formularios rotos, WA no visible, SEO débil.",
        "quien": "Vendedor", "mrr_acum": "Pago único",
        "senal": "Baja tasa de conversión desde tráfico orgánico o Google My Business incompleto."
    },
    {
        "nivel": "Nivel 3A · Ejecución", "producto": "Seguimiento Comercial", "precio": "$3,000 / sesión",
        "cuando": "Cuando el equipo de admisiones del cliente tiene problemas de ejecución post-capacitación.",
        "quien": "Vendedor", "mrr_acum": "Variable",
        "senal": "Cliente con sistema armado pero asesor de admisiones con bajo cierre."
    },
    {
        "nivel": "Nivel 3B · Entrenamiento", "producto": "Taller 'Inscripción es Cuestión de Ventas'", "precio": "$8,500 / persona",
        "cuando": "Cuando el diagnóstico detecta baja tasa de cierre en la cita.",
        "quien": "Vendedor", "mrr_acum": "Pago único",
        "senal": "Cliente cuya cita→inscrito es <40% y tiene equipo de admisiones nuevo o sin entrenamiento."
    },
    {
        "nivel": "Suite Completa · Cliente maduro", "producto": "CRM Avanzado + Dirección Avanzada + Campañas Avanzado",
        "precio": "$60,000+ / mes",
        "cuando": "Mes 6-12 cuando el cliente tiene resultados y quiere escalar.",
        "quien": "Director + Vendedor", "mrr_acum": "$60,000+/mes",
        "senal": "Cliente con >6 meses activo, NPS alto, pide escalar a más niveles o campus."
    }
]

MRR = [
    ("Solo CRM Básico", "50", "$12,000", "$600,000", "$7,200,000", "Entrada mínima. No es el objetivo."),
    ("CRM + Dirección", "50", "$18,000", "$900,000", "$10,800,000", "Primer upsell natural. Meta en año 1."),
    ("CRM + Dir. + Campañas", "50", "$28,000", "$1,400,000", "$16,800,000", "Segundo upsell. Meta en año 2."),
    ("★ META 100 clientes (ticket prom.)", "100", "$18,000", "$1,800,000", "$21,600,000", "PLAN ACTIVO 2026"),
    ("Suite completa (10 clientes)", "10", "$60,000", "$600,000", "$7,200,000", "Prueba del modelo premium."),
    ("Suite completa (50 clientes)", "50", "$60,000", "$3,000,000", "$36,000,000", "Objetivo escala alta."),
]

def build_expansion(wb):
    ws = wb.create_sheet("📈 Ruta de Expansión")
    setup_sheet(ws, BLUE_MID)

    ws["A1"] = "RUTA DE EXPANSIÓN DEL CLIENTE — SUPERLEADS 2026"
    ws.merge_cells("A1:G1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "El cliente no compra más servicios — activa más motor de su rueda. Cada nivel resuelve un cuello más profundo."
    ws.merge_cells("A2:G2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(size=9, color=MUTED)
    ws["A2"].alignment = align(h="left")

    spacer(ws, 3, 7)
    section_title(ws, 4, "CATÁLOGO DE SERVICIOS — NIVELES DE EXPANSIÓN", 7)
    header_row(ws, 5, ["Nivel", "Producto / Servicio", "Precio", "Cuándo proponer", "Quién propone", "MRR acumulado", "Señal de activación"])
    set_height(ws, 5, 22)

    for r, e in enumerate(EXPANSION, 6):
        bg = LIGHT_ROW if r % 2 == 0 else ALT_ROW
        data_row(ws, r, [e["nivel"], e["producto"], e["precio"], e["cuando"], e["quien"], e["mrr_acum"], e["senal"]], bg=bg)
        ws.cell(row=r, column=2).font = font(bold=True, color=WHITE, size=9)
        ws.cell(row=r, column=3).font = font(bold=True, color=GREEN, size=9)
        set_height(ws, r, 55)

    spacer(ws, 15, 7)
    section_title(ws, 16, "PROYECCIÓN DE MRR — ESCENARIOS", 7)
    header_row(ws, 17, ["Escenario", "Clientes", "MRR prom./cliente", "MRR total", "Ingreso anual", "Nota", ""])

    for r, row in enumerate(MRR, 18):
        bg = "003300" if "★" in row[0] else (LIGHT_ROW if r % 2 == 0 else ALT_ROW)
        data_row(ws, r, list(row) + [""], bg=bg)
        if "★" in row[0]:
            for c in range(1, 7):
                ws.cell(row=r, column=c).font = font(bold=True, color=GREEN, size=10)
        set_height(ws, r, 22)

    set_widths(ws, [26, 48, 26, 38, 28, 24, 24])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4: RITMO OPERATIVO
# ══════════════════════════════════════════════════════════════════════════════

DAILY = [
    ("Apertura", "30 min", "Revisar CRM: leads nuevos + stale alerts. Responder mensajes pendientes.", "Setter / Vendedor", "CRM SuperLeads", "Inbox limpio. Leads del día asignados."),
    ("Seguimientos", "60 min", "Ejecutar cadencias activas (WA, llamada, email). Prioridad: Negociación → Propuesta → Cita agendada → Lead.", "Setter / Vendedor", "CRM + WA", "Todas las oportunidades prioritarias tocadas."),
    ("Llamadas y cierres", "60–90 min", "Llamadas para agendar diagnósticos y cerrar negociaciones abiertas.", "Vendedor + Director", "Teléfono + Calendario", "Citas agendadas o decisiones avanzadas."),
    ("Diagnósticos", "Variable", "Sesiones de Rayos X con prospectos. Preparación de propuestas.", "Vendedor", "App Rayos X + Drive", "PDF de diagnóstico o propuesta lista."),
    ("Cierre de día", "15 min", "Actualizar CRM. Dejar siguiente paso con fecha en cada oportunidad tocada.", "Todo el equipo", "CRM SuperLeads", "Cero oportunidades sin fecha de next step."),
]

WEEKLY = [
    ("Lunes 9:00 AM", "Pipeline Review", "1. Stale opps · 2. Prioridades de la semana · 3. Citas confirmadas · 4. Bloqueos", "Director + Vendedores", "30 min", "Top 5 oportunidades a mover. Bloqueos asignados."),
    ("Miércoles 10:00 AM", "Revisión Diagnósticos y Propuestas", "1. Diagnósticos de la semana · 2. Propuestas por enviar · 3. Objeciones activas · 4. Negociaciones estancadas", "Director + Vendedores", "45 min", "Propuestas revisadas. Objeciones con respuesta asignada."),
    ("Viernes 5:00 PM", "Cierre de semana + métricas", "1. KPIs de la semana · 2. Qué funcionó / qué no · 3. Prioridades del lunes", "Todo el equipo", "30 min", "Reporte semanal actualizado. Plan del lunes claro."),
    ("Quincenal (variable)", "Revisión Estratégica con el cliente", "1. Avances vs. metas · 2. Cuello siguiente · 3. Upsell si aplica", "Vendedor + Cliente", "60–90 min", "Minuta enviada al cliente. Upsell identificado si aplica."),
]

MONTHLY_Q = [
    ("Mensual", "Revisión de KPIs globales", "Leads · CPL · % contactación · % show · % cierre · MRR nuevo · Churn · Tickets stale", "Director + Liderazgo", "60 min", "Dashboard actualizado. Ajuste de inversión en ads si CPL fuera de meta."),
    ("Mensual", "Sesión técnica CRM con clientes", "Salud de la Base: datos limpios, automatizaciones activas, embudo correcto.", "Conexión", "30 min/c.", "Clientes sin riesgo de churn. Upsell detectado."),
    ("Trimestral", "Revisión del Plan Comercial", "Avance a Meta 3 (100 clientes) · Ajuste estrategia de atracción · Nuevos consultores necesarios", "Director + Equipo", "90 min", "Plan ajustado para el siguiente trimestre."),
    ("Trimestral", "Calibración de precios", "¿Los precios reflejan el valor entregado? ¿Hay productos a lanzar o retirar?", "Director", "45 min", "Precios actualizados. Comunicado a clientes con 30 días de anticipación."),
]

def build_ritmo(wb):
    ws = wb.create_sheet("🗓️ Ritmo Operativo")
    setup_sheet(ws, "0039C8")

    ws["A1"] = "RITMO OPERATIVO — MÉTODO SUPERLEADS 2026"
    ws.merge_cells("A1:F1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)
    ws["A2"] = "Sin un ritmo fijo, la operación depende de la memoria de las personas, no del sistema."
    ws.merge_cells("A2:F2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(size=9, color=MUTED)
    ws["A2"].alignment = align(h="left")

    row = 4
    # ─ Diaria
    section_title(ws, row, "RUTINA DIARIA — Todos los días hábiles", 6); row += 1
    header_row(ws, row, ["Bloque", "Duración", "Actividad", "Quién", "Herramienta", "Output esperado"]); row += 1
    for d in DAILY:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, d, bg=bg)
        ws.cell(row=row, column=1).font = font(bold=True, color=GREEN, size=9)
        set_height(ws, row, 42)
        row += 1

    spacer(ws, row, 6); row += 1

    # ─ Semanal
    section_title(ws, row, "RITMO SEMANAL — Reuniones y revisiones fijas", 6); row += 1
    header_row(ws, row, ["Día / Hora", "Reunión", "Agenda fija", "Quiénes", "Duración", "Decisión esperada"]); row += 1
    for d in WEEKLY:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, d, bg=bg)
        ws.cell(row=row, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=row, column=2).font = font(bold=True, color=WHITE, size=9)
        set_height(ws, row, 50)
        row += 1

    spacer(ws, row, 6); row += 1

    # ─ Mensual / Trimestral
    section_title(ws, row, "RITMO MENSUAL Y TRIMESTRAL", 6); row += 1
    header_row(ws, row, ["Frecuencia", "Reunión", "Agenda", "Quiénes", "Duración", "Output"]); row += 1
    for d in MONTHLY_Q:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, d, bg=bg)
        ws.cell(row=row, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=row, column=2).font = font(bold=True, color=WHITE, size=9)
        set_height(ws, row, 44)
        row += 1

    set_widths(ws, [22, 36, 60, 28, 16, 48])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 5: OBJECIONES
# ══════════════════════════════════════════════════════════════════════════════

OBJECIONES = [
    {
        "titulo": "«Está caro»",
        "subtitulo": "No ve claramente el valor vs. lo que pierde sin sistema.",
        "respuesta": '"Entiendo. ¿Me permites mostrarte el cálculo que hicimos en el diagnóstico?"\n\nSi no inscribes [X] alumnos más este ciclo, son $[impacto] que no entran. El sistema cuesta $[precio]. ¿El riesgo es el precio o seguir perdiendo esos alumnos?',
        "argumento": "El precio solo parece alto cuando el costo de no actuar no está claro. Siempre cuantificar: (meta – actual) × ticket anual.",
        "si_persiste": '"¿Qué parte del valor no quedó clara?" — No bajar precio aún.',
        "regla": "Nunca bajar precio de entrada. Siempre regresar al impacto económico del diagnóstico."
    },
    {
        "titulo": "«No tengo presupuesto»",
        "subtitulo": "O es real (caja) o es una forma de decir 'no vi el valor'.",
        "respuesta": '"¿El presupuesto no existe o está en otra partida?"\n\nSi es real: "¿Cuándo se define el presupuesto del siguiente ciclo? Podemos agendar para entonces."\n\nSi es excusa: regresar al impacto económico del diagnóstico.',
        "argumento": "Un cliente sin presupuesto hoy puede tenerlo en 60 días. Vale más dejarlo bien posicionado que descartarlo.",
        "si_persiste": "Ofrecer arrancar con CRM Básico ($12k) como entrada mínima de sistema.",
        "regla": "Nunca descartar por presupuesto sin conocer el ciclo de decisión."
    },
    {
        "titulo": "«Déjame pensarlo»",
        "subtitulo": "No hay urgencia clara o falta un decisor involucrado.",
        "respuesta": '"Claro. ¿Qué parte necesitas pensar? ¿Es el precio, el alcance o necesitas validarlo con alguien más?"\n\nSi falta un decisor: "¿Tendría sentido que esa persona esté en la próxima sesión?"',
        "argumento": "Esta objeción siempre esconde algo más específico. Hay que encontrar el bloqueo real, no aceptarla y esperar.",
        "si_persiste": "Poner fecha: '¿El [día] podemos retomar? Así no perdemos el momentum.'",
        "regla": "Nunca dejar esta objeción sin fecha concreta de seguimiento."
    },
    {
        "titulo": "«Ya trabajamos con alguien»",
        "subtitulo": "Tienen algo, pero probablemente sin sistema integrado.",
        "respuesta": '"Perfecto. ¿Están contentos con los resultados? ¿Tienen visibilidad por etapa del embudo?"\n\nSuperLeads no compite con una agencia. Somos el sistema que hace que lo que ya tienen funcione mejor.',
        "argumento": "La mayoría de los 'ya trabajamos con alguien' tienen CRM separado de campañas, separado de seguimiento. No hay sistema — hay piezas sueltas.",
        "si_persiste": "Mostrar el flywheel. Preguntar en cuál etapa están más débiles.",
        "regla": "No competir — posicionarse como el sistema que ordena todo lo que ya tienen."
    },
    {
        "titulo": "«No es el momento»",
        "subtitulo": "El ciclo de inscripciones ya pasó o la urgencia no es visible.",
        "respuesta": '"Entiendo. ¿Cuándo es su próximo ciclo de inscripciones?"\n\nEl mejor momento para instalar el sistema es ANTES del ciclo, no durante. Los que instalan en julio llegan a agosto con el sistema listo.',
        "argumento": "Hay dos momentos malos para instalar: durante el ciclo (sin tiempo) y después del ciclo (sin urgencia). El único momento bueno es ahora.",
        "si_persiste": "Dejar una fecha de seguimiento concreta. No desaparecer.",
        "regla": "Crear urgencia con el ciclo próximo. Nunca dejarlo sin fecha."
    },
    {
        "titulo": "«¿Me pueden garantizar resultados?»",
        "subtitulo": "Busca certeza antes de invertir. Es una señal de compra.",
        "respuesta": '"No garantizamos inscritos — eso depende de tu institución, tu equipo y tu mercado. Lo que sí garantizamos es que vas a tener el sistema que hace posible medirlos, mejorarlos y no perder los que ya tienes."\n\nSin sistema, no hay forma de saber qué falla ni de mejorarlo.',
        "argumento": "Una garantía de inscritos sería irresponsable. Lo que SÍ es garantizable: tener el sistema correcto, medible y operable. Eso es lo que vendemos.",
        "si_persiste": "Mostrar caso similar con métricas reales de mejora de conversión.",
        "regla": "Nunca prometer inscritos. Sí prometer sistema + medición + mejora continua."
    }
]

def build_objeciones(wb):
    ws = wb.create_sheet("💬 Objeciones")
    setup_sheet(ws, "5A00AA")

    ws["A1"] = "MANEJO DE OBJECIONES — MÉTODO SUPERLEADS 2026"
    ws.merge_cells("A1:F1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "Regla: nunca respondas desde inseguridad ni ofrezcas descuento de entrada. Siempre regresa a sistema, impacto económico y costo de seguir igual."
    ws.merge_cells("A2:F2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(bold=True, size=9, color="FFCC44")
    ws["A2"].alignment = align(h="left")

    spacer(ws, 3, 6)
    header_row(ws, 4, ["Objeción", "Por qué surge", "Respuesta recomendada", "Argumento de fondo", "Si persiste…", "Regla de oro"])
    set_height(ws, 4, 22)

    for r, o in enumerate(OBJECIONES, 5):
        bg = LIGHT_ROW if r % 2 == 1 else ALT_ROW
        data_row(ws, r, [o["titulo"], o["subtitulo"], o["respuesta"], o["argumento"], o["si_persiste"], o["regla"]], bg=bg)
        ws.cell(row=r, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=r, column=6).font = font(bold=True, color="FFCC44", size=9, italic=True)
        set_height(ws, r, 90)

    set_widths(ws, [28, 36, 58, 50, 38, 48])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 6: SCORECARD
# ══════════════════════════════════════════════════════════════════════════════

SCORECARD = {
    "🎯 Director": {
        "badge": "Liderazgo · Cierre · Pipeline",
        "metrics": [
            ("Cierres propios (Año 1)", "Pagadas en CRM atribuidas al Director", "2 cierres/sem", "4→2/mes"),
            ("Pipeline Review realizado (lunes 9AM)", "Meeting completado / programado", "1/sem", "4/mes"),
            ("Propuestas revisadas antes de enviar", "Flag 'revisado por Director' en CRM", "100%", "100%"),
            ("Descuentos aprobados (con justificación)", "Descuentos registrados vs. MRR cerrado", "Solo si justificado", "Max 10% MXN cerrado"),
            ("Tiempo promedio Lead → Pagada", "CRM: días entre etapas promedio", "—", "<45 días"),
            ("Pipeline ponderado creciendo", "Delta mensual del pipeline ponderado", "—", "+$500k/mes"),
        ]
    },
    "💼 Vendedor": {
        "badge": "Citas · Pipeline · Cierres",
        "metrics": [
            ("Citas asistidas conducidas", "Citas en etapa 'Asistida' por vendedor", "4 citas", "16/mes"),
            ("Cierres (ventas pagadas)", "Etapa Pagada atribuida al vendedor", "—", "≥2.1/mes"),
            ("Propuestas enviadas en ≤7 días post-cita", "Fecha envío en CRM vs. fecha de cita", "100%", "100%"),
            ("Tasa de avance a Negociación", "Negociaciones / Propuestas enviadas", "—", ">50%"),
            ("CRM actualizado al cierre del día", "Next step con fecha en cada opp", "100%", "100%"),
            ("Oportunidades stale en su pipeline", "% opps sin actividad en 7 días", "<10%", "<10%"),
            ("Upsells detectados y documentados", "Registro en CRM de oportunidad de expansión", "—", "1+/mes"),
        ]
    },
    "📞 Setter": {
        "badge": "Agendamiento · Calificación · Show Rate",
        "metrics": [
            ("Citas nuevas agendadas", "CRM: citas creadas y confirmadas", "5/día", "20 calificadas"),
            ("Show rate de diagnósticos", "Asistidos / Agendados por el Setter", "—", "≥80%"),
            ("Leads respondidos en <3 horas", "CRM: tiempo entre entrada y primer toque", "100%", "100%"),
            ("% Contactación (respondieron)", "Contactados / Leads del período", ">40%", ">40%"),
            ("Citas confirmadas con los 3 requisitos", "Checklist en CRM antes de la cita", "100%", "100%"),
            ("Bono de 100 citas activado", "Citas agendadas en el mes vs. umbral", "—", "≥100 citas"),
        ]
    },
    "🎤 Speaker": {
        "badge": "Eventos · Leads · Contenido",
        "metrics": [
            ("Eventos realizados (webinar, CIMEDU, desayuno)", "Eventos completados en el mes", "—", "1 evento/mes"),
            ("Leads generados por evento", "Leads que entran al funnel desde evento", "—", "≥20 leads"),
            ("Asistentes por evento", "Registro de asistentes confirmados", "—", "≥20 asistentes"),
            ("Tasa de conversión evento → cita", "Citas agendadas / asistentes al evento", "—", ">15%"),
        ],
        "nota": "Rol activo desde Mes 30 (Año 2.5). Si <15 leads/evento × 3 eventos → revisión de contenido."
    },
    "🔗 Conexión": {
        "badge": "Operaciones · CRM · Handoffs",
        "metrics": [
            ("Trámites completados en SLA (<24h)", "Alertas en CRM cerradas en tiempo", "100%", "0 vencidos"),
            ("Clientes con salud de Base >80%", "Checklist Base en CRM del cliente", "—", ">80% portafolio"),
            ("Tickets de soporte resueltos en <48h", "CRM: tiempo apertura → cierre", ">90%", ">90%"),
            ("Upsells identificados y reportados", "Pipeline oportunidades de expansión", "—", "2+/mes"),
            ("Clientes en riesgo de churn detectados", "Reporte mensual de salud de cartera", "—", "100% detectados"),
            ("Churn rate del portafolio", "Cancelaciones / clientes activos", "—", "<5%/mes"),
            ("Sesiones técnicas mensuales realizadas", "Sesiones completadas / clientes activos", "—", "100% clientes"),
        ],
        "nota": "1 Conexión por cada 5 vendedores activos. Rol activo desde Mes 7."
    }
}

def build_scorecard(wb):
    ws = wb.create_sheet("📊 Scorecard")
    setup_sheet(ws, "FF8800")

    ws["A1"] = "SCORECARD DE ROLES — MÉTODO SUPERLEADS 2026"
    ws.merge_cells("A1:F1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "Sin métricas por rol el coaching es opinión — con métricas es datos. Cada persona sabe exactamente qué se espera de ella."
    ws.merge_cells("A2:F2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(size=9, color=MUTED)
    ws["A2"].alignment = align(h="left")

    row = 4
    for role_name, data in SCORECARD.items():
        section_title(ws, row, f"{role_name}  ·  {data['badge']}", 6); set_height(ws, row, 24); row += 1
        header_row(ws, row, ["Métrica", "Cómo se mide", "Meta Semanal / Diaria", "Meta Mensual", "🟢", ""]); row += 1

        for m in data["metrics"]:
            bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
            data_row(ws, row, [m[0], m[1], m[2], m[3], "🟢", ""], bg=bg)
            ws.cell(row=row, column=3).font = font(bold=True, color=GREEN, size=9)
            ws.cell(row=row, column=4).font = font(bold=True, color=GREEN, size=9)
            set_height(ws, row, 22)
            row += 1

        if "nota" in data:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1, value=f"ℹ  {data['nota']}")
            c.fill = fill(NAVY_DEEP)
            c.font = font(size=8, color=MUTED, italic=True)
            c.alignment = align(h="left")
            set_height(ws, row, 18)
            row += 1

        spacer(ws, row, 6); row += 1

    set_widths(ws, [46, 44, 26, 26, 6, 6])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 7: EQUIPO
# ══════════════════════════════════════════════════════════════════════════════

ROLES_COMP = [
    {
        "role": "🎯 Director", "desde": "Mes 1", "funcion": "Lidera el equipo, vende en Año 1, cierra negociaciones complejas y valida el funnel.",
        "sueldo": "$75k → $90k → $105k → $120k / mes", "comision": "10% de sus ventas propias",
        "bono": "—", "social": "20% sobre base", "incremento": "+$15,000/año"
    },
    {
        "role": "💼 Vendedor", "desde": "Mes 3", "funcion": "Gestiona el pipeline completo de citas a cierres. Meta: ≥2.1 ventas/mes.",
        "sueldo": "$15k → $20k → $25k → $30k / mes", "comision": "15% del contrato cerrado",
        "bono": "$1,000 por cada cierre", "social": "—", "incremento": "+$5,000/año"
    },
    {
        "role": "📞 Setter", "desde": "Mes 12", "funcion": "Agenda citas para los vendedores. Califica leads y confirma asistencia.",
        "sueldo": "$10,000 fijo / mes", "comision": "$2,000 por venta cerrada de sus citas",
        "bono": "$5,000 si agenda ≥100 citas/mes", "social": "20% sobre base", "incremento": "—"
    },
    {
        "role": "🎤 Speaker", "desde": "Mes 30", "funcion": "Vendedor con perfil de ponente. Da webinars, CIMEDU y desayunos para captar leads.",
        "sueldo": "$10,000 fijo / mes", "comision": "Sin comisión de ventas",
        "bono": "—", "social": "20% sobre base", "incremento": "—"
    },
    {
        "role": "🔗 Conexión", "desde": "Mes 7 (1 por c/5 vendedores)", "funcion": "Agiliza trámites, coordinación, CRM y handoffs. Reduce fricción operativa.",
        "sueldo": "$10,000 fijo / mes", "comision": "Sin comisiones de venta",
        "bono": "—", "social": "20% sobre base", "incremento": "—"
    }
]

HIRING = [
    ("M1",  "May-26", "🎯 Director (arranque)",                       "Valida el funnel y vende solo",                "—",  "—",   "—"),
    ("M3",  "Jul-26", "💼 Vendedor #1",                               "Funnel probado, 20 citas/mes confirmadas",     "1",  "—",   "—"),
    ("M6",  "Oct-26", "💼 Vendedor #2",                               "Vendedor #1 llena su agenda completa",         "2",  "—",   "—"),
    ("M7",  "Nov-26", "🔗 Conexión #1",                               "2 vendedores → fricción operativa real",       "2",  "—",   "1"),
    ("M12", "Abr-27", "📞 Setter #1 + 💼 Vendedor #3",               "Escala a 60+ citas/mes, se necesita agendar",  "3",  "1",   "1"),
    ("M18", "Oct-27", "💼 Vendedor #4",                               "Meta 20+ cierres/mes activada",                "4",  "1",   "1"),
    ("M24", "Abr-28", "💼 Vendedor #5",                               "Meta 32+ cierres/mes, inicio Año 3",           "5",  "1",   "1"),
    ("M30", "Oct-28", "🎤 Speaker #1 + 💼 Vendedor #6",              "Año 2.5: escala eventos y contenido",          "6",  "1",   "1"),
    ("M36", "Abr-29", "💼 Vendedor #7",                               "Consolidar Año 3",                             "7",  "1",   "1"),
    ("M42", "Oct-29", "💼 Vendedor #8",                               "Equipo maduro, 8 vendedores",                  "8",  "1",   "1"),
]

DESVINCUL = [
    ("<1 venta/mes × 2 meses (Vendedor/Director)", "2 meses <50% meta", "PIP 30 días → desvinculación", "60 días", "Director"),
    ("No contactar leads en <3h (>20% del mes)", "<80% cumplimiento", "Advertencia → PIP → desvinculación", "30 días", "Director"),
    ("Setter: show rate <60% por 2 meses seguidos", "<60% por 2 meses", "Capacitación → PIP → desvinculación", "45 días", "Director"),
    ("Speaker: <15 leads/evento × 3 eventos", "Promedio <15 leads", "Revisión de contenido → PIP", "60 días", "Director"),
    ("Conexión: SLA >48h de forma habitual", "≥3 incidencias/mes", "Advertencia escrita → PIP", "30 días", "Director"),
    ("Falsificación de actividad en CRM", "Una sola vez comprobada", "Desvinculación inmediata", "Inmediato", "Director"),
    ("Actitud tóxica / conflicto con equipo", "2 reportes formales", "Mediación → PIP → desvinculación", "30–45 días", "Director"),
]

def build_equipo(wb):
    ws = wb.create_sheet("👥 Equipo")
    setup_sheet(ws, "CC0044")

    ws["A1"] = "ORGANIGRAMA COMERCIAL — MÉTODO SUPERLEADS 2026"
    ws.merge_cells("A1:H1")
    ws["A1"].fill = fill(NAVY_DEEP)
    ws["A1"].font = font(bold=True, size=14, color=GREEN)
    ws["A1"].alignment = align(h="center")
    set_height(ws, 1, 32)

    ws["A2"] = "El equipo crece conforme el funnel lo justifica — no antes. Cada rol entra cuando hay un número que lo exige, no una sensación."
    ws.merge_cells("A2:H2")
    ws["A2"].fill = fill(NAVY_DEEP)
    ws["A2"].font = font(size=9, color=MUTED)
    ws["A2"].alignment = align(h="left")

    row = 4
    # ─ Compensaciones
    section_title(ws, row, "ESTRUCTURA DE COMPENSACIÓN POR ROL", 8); row += 1
    header_row(ws, row, ["Rol", "Desde", "Función principal", "Sueldo base", "Comisión", "Bono", "Carga social", "Incremento anual"]); row += 1

    for r in ROLES_COMP:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, [r["role"], r["desde"], r["funcion"], r["sueldo"], r["comision"], r["bono"], r["social"], r["incremento"]], bg=bg)
        ws.cell(row=row, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=row, column=4).font = font(bold=True, color=WHITE, size=9)
        set_height(ws, row, 44)
        row += 1

    spacer(ws, row, 8); row += 1

    # ─ Timeline
    section_title(ws, row, "PLAN DE CONTRATACIÓN — 48 MESES", 8); row += 1
    header_row(ws, row, ["Mes", "Fecha", "Rol que entra", "Disparador de contratación", "Vendedores", "Setters", "Conexión", ""]); row += 1

    for h in HIRING:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, list(h) + [""], bg=bg)
        ws.cell(row=row, column=1).font = font(bold=True, color=GREEN, size=9)
        ws.cell(row=row, column=3).font = font(bold=True, color=WHITE, size=9)
        set_height(ws, row, 22)
        row += 1

    spacer(ws, row, 8); row += 1

    # ─ Desvinculación
    section_title(ws, row, "POLÍTICA DE PERMANENCIA Y DESVINCULACIÓN", 8); row += 1
    ws["A" + str(row)] = "PIP (Plan de Mejora) de 30 días si se incumplen 2+ métricas durante 2 meses consecutivos. Sin mejora al término → desvinculación."
    ws.merge_cells(f"A{row}:H{row}")
    ws["A" + str(row)].fill = fill(NAVY_DEEP)
    ws["A" + str(row)].font = font(size=9, color="FF8080")
    ws["A" + str(row)].alignment = align(h="left")
    row += 1
    header_row(ws, row, ["Causal", "Umbral", "Proceso", "Plazo", "Dueño", "", "", ""]); row += 1

    for d in DESVINCUL:
        bg = LIGHT_ROW if row % 2 == 1 else ALT_ROW
        data_row(ws, row, list(d) + ["", "", ""], bg=bg)
        ws.cell(row=row, column=3).font = font(bold=True, color="FF8080", size=9)
        set_height(ws, row, 22)
        row += 1

    set_widths(ws, [28, 26, 52, 30, 22, 16, 16, 12])


# ══════════════════════════════════════════════════════════════════════════════
# PORTADA
# ══════════════════════════════════════════════════════════════════════════════

def build_cover(wb):
    ws = wb.active
    ws.title = "📋 Portada"
    setup_sheet(ws, GREEN)

    ws.sheet_view.showGridLines = False
    set_widths(ws, [8, 50, 50, 8])

    for r in range(1, 30):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill(NAVY_DEEP)

    # Title block
    ws.merge_cells("B3:C3")
    ws["B3"] = "SUPERLEADS"
    ws["B3"].font = Font(bold=True, size=32, color=GREEN, name="Calibri")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    set_height(ws, 3, 48)

    ws.merge_cells("B4:C4")
    ws["B4"] = "MÉTODO COMERCIAL 2026"
    ws["B4"].font = Font(bold=True, size=18, color=WHITE, name="Calibri")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    set_height(ws, 4, 32)

    ws.merge_cells("B5:C5")
    ws["B5"] = "Manual del Equipo Comercial"
    ws["B5"].font = Font(size=11, color=MUTED, name="Calibri")
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
    set_height(ws, 5, 22)

    # Separator
    ws.merge_cells("B7:C7")
    ws["B7"] = "─" * 60
    ws["B7"].font = Font(size=9, color="0D2A6E", name="Calibri")
    ws["B7"].alignment = Alignment(horizontal="center")
    set_height(ws, 7, 14)

    # Tabs index
    tabs = [
        ("🔄", "Pipeline",             "11 etapas del funnel con dueño, objetivo y tiempos máximos."),
        ("✅", "Definition of Done",   "Criterios obligatorios para avanzar cada oportunidad."),
        ("📈", "Ruta de Expansión",    "Catálogo de servicios, niveles y proyección de MRR."),
        ("🗓️", "Ritmo Operativo",      "Cadencias diarias, semanales, mensuales y trimestrales."),
        ("💬", "Objeciones",           "Las 6 objeciones más frecuentes y cómo responderlas."),
        ("📊", "Scorecard",            "Métricas por rol: Director, Vendedor, Setter, Speaker, Conexión."),
        ("👥", "Equipo",               "Compensaciones, plan de contratación 48 meses y política de permanencia."),
    ]

    ws.merge_cells("B9:C9")
    ws["B9"] = "CONTENIDO"
    ws["B9"].font = Font(bold=True, size=10, color=GREEN, name="Calibri")
    ws["B9"].alignment = Alignment(horizontal="left", vertical="center")
    set_height(ws, 9, 20)

    for i, (icon, tab, desc) in enumerate(tabs):
        r = 10 + i * 2
        ws.cell(row=r, column=2).value = f"  {icon}  {tab}"
        ws.cell(row=r, column=2).font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        ws.cell(row=r, column=2).fill = fill(NAVY_MID if i % 2 == 0 else LIGHT_ROW)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=3).value = desc
        ws.cell(row=r, column=3).font = Font(size=9, color=MUTED, name="Calibri")
        ws.cell(row=r, column=3).fill = fill(NAVY_MID if i % 2 == 0 else LIGHT_ROW)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")
        set_height(ws, r, 22)

    # Footer
    ws.merge_cells("B26:C26")
    ws["B26"] = "Versión 2026 · Confidencial · Uso interno SuperLeads"
    ws["B26"].font = Font(size=8, color="334466", name="Calibri", italic=True)
    ws["B26"].alignment = Alignment(horizontal="center")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    build_cover(wb)
    build_pipeline(wb)
    build_dod(wb)
    build_expansion(wb)
    build_ritmo(wb)
    build_objeciones(wb)
    build_scorecard(wb)
    build_equipo(wb)

    output = "/Users/constanza/Desktop/Método_Comercial_SuperLeads_2026.xlsx"
    wb.save(output)
    print(f"✅  Guardado en: {output}")

if __name__ == "__main__":
    main()
